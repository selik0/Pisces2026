"""读取 GameProto 约定的 Excel 配置表布局。"""
from dataclasses import dataclass
import re
from pathlib import Path

SUPPORTED = {"uint", "int", "bool", "float", "long", "double", "string", "bytes", "byte", "short", "ushort", "ulong", "sbyte"}
RANGE = {"ALL", "CLIENT", "SERVER"}
IDENTIFIER = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

@dataclass
class Field:
    name: str
    type_name: str
    source: str
    export: str
    description: str
    column: int

@dataclass
class ConfigSheet:
    name: str
    class_name: str
    local_class: bool
    key_fields: list
    group_fields: list
    fields: list
    rows: list

class ExcelError(ValueError):
    pass

def _text(value):
    return "" if value is None else str(value).strip()

def _parts(value, label, path, sheet):
    result = [x.strip() for x in _text(value).split(".") if x.strip()]
    if len(result) > 3:
        raise ExcelError(f"{path}，Sheet={sheet}，{label}最多支持3个字段")
    return result

def _validate_type(type_name, path, sheet):
    t = type_name.replace(" ", "")
    if t in SUPPORTED:
        return t
    if t.endswith("[]"):
        element = t[:-2]
        if element not in SUPPORTED:
            raise ExcelError(f"{path}，Sheet={sheet}，不支持数组元素类型：{element}")
        return t
    if t.startswith("List<") and t.endswith(">"):
        element = t[5:-1]
        if element not in SUPPORTED:
            raise ExcelError(f"{path}，Sheet={sheet}，不支持 List 元素类型：{element}")
        return t
    raise ExcelError(f"{path}，Sheet={sheet}，不支持字段类型：{type_name}")

def load_workbook(path, sheet_filter=None, data_start_row=7):
    try:
        from openpyxl import load_workbook as openpyxl_load
    except ImportError as exc:
        raise ExcelError("缺少 openpyxl，请执行：python -m pip install openpyxl") from exc
    file_path = Path(path)
    if not file_path.is_file():
        raise ExcelError(f"Excel 文件不存在：{file_path}")
    workbook = openpyxl_load(str(file_path), data_only=True, read_only=True)
    result = []
    classes = {}
    for worksheet in workbook.worksheets:
        if sheet_filter and worksheet.title not in sheet_filter:
            continue
        first = [_text(worksheet.cell(1, col).value) for col in range(1, worksheet.max_column + 1)]
        class_name = first[0] if first else ""
        # Sheet1 是模板说明页，不是可导出的配置表。
        if not class_name or not IDENTIFIER.match(class_name.split("（")[0]):
            continue
        class_name = class_name.split("（")[0].strip()
        if not IDENTIFIER.match(class_name):
            raise ExcelError(f"{file_path}，Sheet={worksheet.title}，类名非法：{class_name}")
        local_class = first[1].upper() == "TRUE" if len(first) > 1 else False
        key_fields = _parts(first[2] if len(first) > 2 else "", "索引字段", file_path, worksheet.title)
        group_fields = _parts(first[3] if len(first) > 3 else "", "分类字段", file_path, worksheet.title)
        fields = []
        seen = set()
        for col in range(2, worksheet.max_column + 1):
            desc = _text(worksheet.cell(2, col).value)
            name = _text(worksheet.cell(3, col).value)
            type_name = _text(worksheet.cell(4, col).value).replace(" ", "")
            source = _text(worksheet.cell(5, col).value)
            export = _text(worksheet.cell(6, col).value) or "All"
            if not any((desc, name, type_name, source, export)):
                continue
            location = f"{file_path}，Sheet={worksheet.title}，Excel行=3，列={col}，字段={name or '<空>'}"
            if not name or not type_name:
                raise ExcelError(f"{location}，字段名和类型不能为空")
            if not IDENTIFIER.match(name):
                raise ExcelError(f"{location}，字段名不是合法 C# 标识符")
            type_name = _validate_type(type_name, location, worksheet.title)
            if export.upper() not in RANGE:
                raise ExcelError(f"{location}，导出范围必须为 All、Client 或 Server")
            identity = (name, source, export.upper(), type_name)
            if identity in seen and source not in {"List", "ClassStructList", "ClassStructArray"}:
                raise ExcelError(f"{location}，重复字段无法安全区分")
            # List/结构体容器中的重复列由父类型和列顺序区分，保留供后续嵌套类型生成。
            seen.add(identity)
            fields.append(Field(name, type_name, source, export.title(), desc, col))
        if not fields:
            raise ExcelError(f"{file_path}，Sheet={worksheet.title}，没有字段")
        names = {field.name for field in fields}
        for label, indexes in (("索引字段", key_fields), ("分类字段", group_fields)):
            for key in indexes:
                if key not in names:
                    raise ExcelError(f"{file_path}，Sheet={worksheet.title}，{label}不存在：{key}")
        rows = []
        for row_number in range(data_start_row, worksheet.max_row + 1):
            values = [_text(worksheet.cell(row_number, field.column).value) for field in fields]
            if not any(values):
                continue
            rows.append((row_number, values))
        if class_name in classes and not local_class:
            raise ExcelError(f"{file_path}，Sheet={worksheet.title}，配置类名冲突：{class_name}")
        classes[class_name] = worksheet.title
        result.append(ConfigSheet(worksheet.title, class_name, local_class, key_fields, group_fields, fields, rows))
    workbook.close()
    if not result:
        raise ExcelError(f"{file_path} 中没有找到可导出的配置 Sheet")
    return result
